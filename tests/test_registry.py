from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path
from uuid import uuid4

import pandas as pd
from openpyxl import load_workbook

from quant_studio_monitoring.config import WorkspaceConfig
from quant_studio_monitoring.registry import (
    MonitoringMetadata,
    build_input_template_workbook_bytes,
    build_model_bundle_intake_checklist_frame,
    build_reference_baseline_diagnostics_frame,
    discover_model_bundles,
    save_monitoring_metadata,
    summarize_dataset_contract,
    validate_input_contract,
)


def test_discover_model_bundle_reads_contract_and_metadata() -> None:
    workspace = _build_workspace()
    bundle_root = workspace.models_root / "retail_pd_v1"
    bundle_root.mkdir(parents=True)

    _write_bundle_files(bundle_root)

    bundles = discover_model_bundles(workspace)

    assert len(bundles) == 1
    bundle = bundles[0]
    assert bundle.display_name == "Retail PD"
    assert bundle.model_version == "2026.01"
    assert bundle.model_owner == "Model Risk"
    assert bundle.model_type == "logistic_regression"
    assert bundle.target_mode == "binary"
    assert bundle.default_segment_column == "region"
    assert bundle.reference_score_column == "predicted_probability_recommended"
    assert bundle.expected_input_columns == ["as_of_date", "loan_id", "balance", "region"]
    assert bundle.optional_input_columns == ["default_status"]
    assert bundle.readiness_status == "ready_with_warnings"
    assert bundle.compatibility_status == "compatible_with_warnings"
    assert bundle.monitoring_contract_version_status == "missing"
    assert bundle.review_status == "review_incomplete"
    assert "Approval Status" in bundle.review_metadata_gaps
    assert bundle.is_compliant is True


def test_validate_input_contract_flags_missing_required_columns() -> None:
    workspace = _build_workspace()
    bundle_root = workspace.models_root / "retail_pd_v1"
    bundle_root.mkdir(parents=True)
    _write_bundle_files(bundle_root)
    bundle = discover_model_bundles(workspace)[0]

    dataframe = pd.DataFrame(
        {
            "as_of_date": ["2026-01-01"],
            "loan_id": ["L1"],
            "region": ["north"],
            "default_status": [1],
        }
    )
    contract = validate_input_contract(bundle, dataframe, segment_column="region")

    assert contract.missing_required_columns == ["balance"]
    assert contract.missing_optional_columns == []
    assert contract.labels_available is True
    assert contract.resolved_label_column == "default_status"
    assert contract.segment_available is True


def test_validate_input_contract_tracks_missing_optional_schema_columns() -> None:
    workspace = _build_workspace()
    bundle_root = workspace.models_root / "retail_pd_v1"
    bundle_root.mkdir(parents=True)
    _write_bundle_files(bundle_root, include_ignore_column=True)
    bundle = discover_model_bundles(workspace)[0]

    dataframe = pd.DataFrame(
        {
            "as_of_date": ["2026-01-01"],
            "loan_id": ["L1"],
            "balance": [100.0],
            "region": ["north"],
        }
    )
    contract = validate_input_contract(bundle, dataframe)

    assert contract.missing_required_columns == []
    assert contract.missing_optional_columns == ["legacy_text_field", "default_status"]
    assert contract.labels_available is False


def test_save_monitoring_metadata_persists_manifest_values() -> None:
    workspace = _build_workspace()
    bundle_root = workspace.models_root / "retail_pd_v1"
    bundle_root.mkdir(parents=True)
    _write_bundle_files(bundle_root)
    bundle = discover_model_bundles(workspace)[0]

    saved_path = save_monitoring_metadata(
        bundle,
        MonitoringMetadata(
            model_name="Retail PD Monitored",
            model_version="2026.04",
            model_owner="Second Line",
            business_purpose="Ongoing monitoring",
            portfolio_name="Retail",
            segment_name="Region",
            default_segment_column="region",
            approval_status="approved",
            approval_date="2026-04-21",
            monitoring_notes="Managed by the monitoring app.",
        ),
    )
    refreshed_bundle = discover_model_bundles(workspace)[0]

    assert saved_path.exists()
    assert refreshed_bundle.display_name == "Retail PD Monitored"
    assert refreshed_bundle.monitoring_metadata.approval_status == "approved"
    assert refreshed_bundle.metadata_source == "manifest"


def test_summarize_dataset_contract_reports_score_only_warning() -> None:
    workspace = _build_workspace()
    bundle_root = workspace.models_root / "retail_pd_v1"
    bundle_root.mkdir(parents=True)
    _write_bundle_files(bundle_root, include_ignore_column=True)
    bundle = discover_model_bundles(workspace)[0]

    dataframe = pd.DataFrame(
        {
            "as_of_date": ["2026-01-01", "2026-01-02"],
            "loan_id": ["L1", "L2"],
            "balance": [100.0, 200.0],
            "region": ["north", "south"],
        }
    )
    summary = summarize_dataset_contract(bundle, dataframe, segment_column="region")

    assert summary.overall_status == "ready_with_warnings"
    assert summary.score_only_run is True
    assert not summary.hard_failures
    assert any("score-only mode" in warning for warning in summary.warnings)
    assert "legacy_text_field" in summary.column_checks["column_name"].tolist()


def test_build_input_template_workbook_contains_expected_sheets_and_headers() -> None:
    workspace = _build_workspace()
    bundle_root = workspace.models_root / "retail_pd_v1"
    bundle_root.mkdir(parents=True)
    _write_bundle_files(bundle_root, include_ignore_column=True)
    bundle = discover_model_bundles(workspace)[0]

    workbook_bytes = build_input_template_workbook_bytes(bundle, include_example_row=True)
    workbook = load_workbook(filename=BytesIO(workbook_bytes))

    assert workbook.sheetnames == ["input_template", "column_guide", "usage_notes"]
    headers = [cell.value for cell in workbook["input_template"][1]]
    assert headers == [
        "as_of_date",
        "loan_id",
        "balance",
        "region",
        "legacy_text_field",
        "default_status",
    ]


def test_discover_standalone_model_file_as_incomplete_bundle() -> None:
    workspace = _build_workspace()
    model_path = workspace.models_root / "quant_model.joblib"
    model_path.write_bytes(b"placeholder")

    bundles = discover_model_bundles(workspace)

    assert len(bundles) == 1
    bundle = bundles[0]
    assert bundle.display_name == "quant_model"
    assert bundle.bundle_paths.model_path == model_path
    assert bundle.is_compliant is False
    assert "Missing run_config.json." in bundle.issues
    assert any("Standalone model artifact detected" in issue for issue in bundle.issues)


def test_discover_prefers_nested_monitoring_bundle_shape() -> None:
    workspace = _build_workspace()
    parent_root = workspace.models_root / "retail_pd_v1"
    bundle_root = parent_root / "monitoring_bundle"
    parent_root.mkdir(parents=True)
    (parent_root / "quant_model.joblib").write_bytes(b"legacy-placeholder")
    bundle_root.mkdir(parents=True)
    _write_bundle_files(bundle_root)

    bundles = discover_model_bundles(workspace)

    assert len(bundles) == 1
    bundle = bundles[0]
    assert bundle.bundle_id == "retail_pd_v1_monitoring_bundle"
    assert bundle.bundle_paths.root == bundle_root
    assert any(check.code == "legacy_monitoring_bundle_shape" for check in bundle.compatibility_checks)


def test_discover_prefers_sister_project_model_bundle_for_monitoring_shape() -> None:
    workspace = _build_workspace()
    parent_root = workspace.models_root / "retail_pd_v1"
    bundle_root = parent_root / "model_bundle_for_monitoring"
    parent_root.mkdir(parents=True)
    (parent_root / "quant_model.joblib").write_bytes(b"legacy-placeholder")
    bundle_root.mkdir(parents=True)
    _write_bundle_files(bundle_root)

    bundles = discover_model_bundles(workspace)

    assert len(bundles) == 1
    bundle = bundles[0]
    assert bundle.bundle_id == "retail_pd_v1_model_bundle_for_monitoring"
    assert bundle.bundle_paths.root == bundle_root
    assert any(
        check.code == "preferred_model_bundle_for_monitoring_shape"
        for check in bundle.compatibility_checks
    )


def test_reference_baseline_diagnostics_summarize_reference_assets() -> None:
    workspace = _build_workspace()
    bundle_root = workspace.models_root / "retail_pd_v1"
    bundle_root.mkdir(parents=True)
    _write_bundle_files(bundle_root)
    bundle = discover_model_bundles(workspace)[0]

    diagnostics = build_reference_baseline_diagnostics_frame(bundle).set_index("diagnostic")

    assert diagnostics.loc["reference_input_rows", "value"] == 2
    assert diagnostics.loc["reference_score_column", "value"] == "predicted_probability_recommended"


def test_bundle_contract_version_and_intake_checklist_are_reported() -> None:
    workspace = _build_workspace()
    bundle_root = workspace.models_root / "retail_pd_v1"
    bundle_root.mkdir(parents=True)
    _write_bundle_files(bundle_root, monitoring_contract_version="1.0")
    bundle = discover_model_bundles(workspace)[0]

    checklist = build_model_bundle_intake_checklist_frame(bundle).set_index("item")

    assert bundle.monitoring_contract_version == "1.0"
    assert bundle.monitoring_contract_version_status == "supported"
    assert checklist.loc["monitoring_contract_version", "status"] == "pass"
    assert any(
        check.code == "supported_monitoring_bundle_version"
        for check in bundle.compatibility_checks
    )


def test_dataset_guardrails_warn_for_large_shapes() -> None:
    workspace = _build_workspace()
    workspace.performance.large_dataset_warning_rows = 2
    workspace.performance.large_dataset_block_rows = 10
    bundle_root = workspace.models_root / "retail_pd_v1"
    bundle_root.mkdir(parents=True)
    _write_bundle_files(bundle_root)
    bundle = discover_model_bundles(workspace)[0]

    dataframe = pd.DataFrame(
        {
            "as_of_date": ["2026-01-01", "2026-01-02"],
            "loan_id": ["L1", "L2"],
            "balance": [100.0, 200.0],
            "region": ["north", "south"],
        }
    )
    summary = summarize_dataset_contract(
        bundle,
        dataframe,
        performance=workspace.performance,
    )

    assert "warning" in summary.guardrails["status"].tolist()


def _build_workspace() -> WorkspaceConfig:
    tmp_path = _make_test_root("registry")
    workspace = WorkspaceConfig(
        project_root=tmp_path,
        models_root=tmp_path / "models",
        incoming_data_root=tmp_path / "incoming_data",
        thresholds_root=tmp_path / "thresholds",
        runs_root=tmp_path / "runs",
    )
    workspace.ensure_directories()
    return workspace


def _make_test_root(prefix: str) -> Path:
    root = Path.cwd() / ".test_workspace" / f"{prefix}_{uuid4().hex}"
    root.mkdir(parents=True, exist_ok=True)
    return root


def _write_bundle_files(
    bundle_root,
    *,
    include_ignore_column: bool = False,
    monitoring_contract_version: str | None = None,
) -> None:
    column_specs = [
        {"name": "as_of_date", "source_name": "as_of_date", "enabled": True, "role": "date"},
        {"name": "loan_id", "source_name": "loan_id", "enabled": True, "role": "identifier"},
        {"name": "balance", "source_name": "balance", "enabled": True, "role": "feature"},
        {"name": "region", "source_name": "region", "enabled": True, "role": "feature"},
    ]
    if include_ignore_column:
        column_specs.append(
            {
                "name": "legacy_text_field",
                "source_name": "legacy_text_field",
                "enabled": True,
                "role": "ignore",
            }
        )
    column_specs.append(
        {
            "name": "default_status",
            "source_name": "default_status",
            "enabled": True,
            "role": "target_source",
        }
    )
    run_config = {
        "schema": {
            "column_specs": column_specs
        },
        "documentation": {
            "model_name": "Retail PD",
            "model_owner": "Model Risk",
            "business_purpose": "PD monitoring",
            "portfolio_name": "Retail",
        },
        "model": {"model_type": "logistic_regression", "threshold": 0.55},
        "target": {"mode": "binary", "output_column": "default_flag"},
        "diagnostics": {"default_segment_column": "region"},
    }
    monitoring_metadata = {
        "model_version": "2026.01",
        "segment_name": "Region",
    }
    if monitoring_contract_version is not None:
        monitoring_metadata["monitoring_contract_version"] = monitoring_contract_version
    (bundle_root / "run_config.json").write_text(json.dumps(run_config), encoding="utf-8")
    (bundle_root / "monitoring_metadata.json").write_text(
        json.dumps(monitoring_metadata),
        encoding="utf-8",
    )
    (bundle_root / "quant_model.joblib").write_bytes(b"placeholder")
    (bundle_root / "generated_run.py").write_text("print('stub')\n", encoding="utf-8")
    pd.DataFrame(
        {
            "as_of_date": ["2025-01-01", "2025-01-02"],
            "loan_id": ["L1", "L2"],
            "balance": [100.0, 200.0],
            "region": ["north", "south"],
        }
    ).to_csv(bundle_root / "input_snapshot.csv", index=False)
    pd.DataFrame(
        {
            "predicted_probability_recommended": [0.1, 0.2],
            "default_flag": [0, 1],
        }
    ).to_csv(bundle_root / "predictions.csv", index=False)
